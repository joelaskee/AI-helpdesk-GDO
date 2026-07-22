import io
import logging
import os
import time
import uuid

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy import select

from .db import Base, SessionLocal, engine
from .models import Call
from .services import llm, pdv, pipeline, transcription

logging.basicConfig(level=logging.INFO)

AUDIO_DIR = os.environ.get("AUDIO_DIR", "/data/audio")
os.makedirs(AUDIO_DIR, exist_ok=True)

app = FastAPI(title="Helpdesk Call Intelligence")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    # Retry: il DB potrebbe non essere ancora pronto
    for _ in range(30):
        try:
            Base.metadata.create_all(engine)
            return
        except Exception:
            time.sleep(2)
    Base.metadata.create_all(engine)


@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "transcribe_engine": transcription.ENGINE,
        "whisper_model": transcription.WHISPER_MODEL,
        "ollama": llm.check_ollama(),
    }


# ---- Anagrafica PDV ----------------------------------------------------------

# Mappa colonne Excel -> campi tabella (robusta a maiuscole/spazi)
_PDV_COLS = {
    "nome pdv": "codice",
    "intestazione pdv": "intestazione",
    "cognome": "cognome",
    "via indirizzo postale": "indirizzo",
    "città indirizzo postale": "citta",
    "citta indirizzo postale": "citta",
    "tipo pdv": "tipo",
}


@app.get("/api/pdv/status")
def pdv_status():
    return {"count": pdv.count()}


@app.post("/api/pdv/import")
async def pdv_import(file: UploadFile = File(...)):
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel non leggibile: {e}")

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Foglio vuoto")

    header = [str(h or "").strip().lower() for h in rows[0]]
    idx = {}
    for i, h in enumerate(header):
        if h in _PDV_COLS:
            idx[_PDV_COLS[h]] = i
    if "codice" not in idx:
        raise HTTPException(400, "Colonna 'Nome Pdv' non trovata nel foglio")

    parsed = []
    for r in rows[1:]:
        if not r or r[idx["codice"]] in (None, ""):
            continue
        rec = {}
        for field, i in idx.items():
            val = r[i] if i < len(r) else None
            rec[field] = str(val).strip() if val not in (None, "") else None
        # il codice può arrivare come numero: normalizza senza decimali
        code = rec.get("codice") or ""
        if code.endswith(".0"):
            code = code[:-2]
        rec["codice"] = code
        parsed.append(rec)

    n = pdv.import_rows(parsed)
    return {"imported": n, "total": pdv.count()}


# ---- Chiamate ----------------------------------------------------------------


@app.post("/api/calls")
async def create_call(audio: UploadFile = File(...)):
    ext = os.path.splitext(audio.filename or "")[1] or ".webm"
    call_id = str(uuid.uuid4())
    path = os.path.join(AUDIO_DIR, f"{call_id}{ext}")
    with open(path, "wb") as f:
        f.write(await audio.read())

    with SessionLocal() as db:
        call = Call(
            id=call_id,
            original_filename=audio.filename,
            audio_path=path,
            status="uploaded",
        )
        db.add(call)
        db.commit()

    pipeline.enqueue(call_id)
    return {"id": call_id, "status": "uploaded"}


@app.get("/api/calls")
def list_calls():
    with SessionLocal() as db:
        calls = db.scalars(select(Call).order_by(Call.created_at.desc())).all()
        return [
            {
                "id": c.id,
                "created_at": c.created_at.isoformat() if c.created_at else None,
                "original_filename": c.original_filename,
                "duration_sec": c.duration_sec,
                "status": c.status,
                "engine": c.engine,
                "has_ticket": c.ticket is not None,
                "completeness_score": (c.completeness or {}).get("score"),
            }
            for c in calls
        ]


def _get_call(call_id: str) -> Call:
    with SessionLocal() as db:
        call = db.get(Call, call_id)
        if call is None:
            raise HTTPException(404, "Chiamata non trovata")
        return call


@app.get("/api/calls/{call_id}")
def get_call(call_id: str):
    return _get_call(call_id).to_dict()


@app.delete("/api/calls/{call_id}")
def delete_call(call_id: str):
    with SessionLocal() as db:
        call = db.get(Call, call_id)
        if call is None:
            raise HTTPException(404, "Chiamata non trovata")
        if call.audio_path and os.path.exists(call.audio_path):
            os.remove(call.audio_path)
        db.delete(call)
        db.commit()
    return {"deleted": call_id}


@app.get("/api/calls/{call_id}/audio")
def get_audio(call_id: str):
    call = _get_call(call_id)
    if not call.audio_path or not os.path.exists(call.audio_path):
        raise HTTPException(404, "Audio non trovato")
    return FileResponse(call.audio_path, filename=call.original_filename or "audio")


@app.put("/api/calls/{call_id}/ticket")
def save_ticket(call_id: str, ticket: dict):
    with SessionLocal() as db:
        call = db.get(Call, call_id)
        if call is None:
            raise HTTPException(404, "Chiamata non trovata")
        call.ticket = ticket
        call.completeness = None  # il ticket è cambiato: invalida l'analisi precedente
        db.commit()
    return {"id": call_id, "ticket": ticket}


@app.post("/api/calls/{call_id}/completeness")
def completeness(call_id: str):
    call = _get_call(call_id)
    if not call.transcript_text:
        raise HTTPException(409, "Trascrizione non ancora disponibile")
    if not call.ticket:
        raise HTTPException(409, "Ticket CRM non ancora compilato")
    pipeline.run_completeness(call_id)
    return _get_call(call_id).to_dict()


@app.post("/api/calls/{call_id}/reprocess")
def reprocess(call_id: str):
    _get_call(call_id)
    pipeline.enqueue(call_id)
    return {"id": call_id, "status": "requeued"}


@app.post("/api/calls/{call_id}/identify-store")
def identify_store(call_id: str):
    """Ri-esegue solo il riconoscimento PDV (senza ritrascrivere)."""
    call = _get_call(call_id)
    if not call.transcript_text:
        raise HTTPException(409, "Trascrizione non ancora disponibile")
    result = pdv.identify(call.transcript_text)
    with SessionLocal() as db:
        c = db.get(Call, call_id)
        c.store_match = result
        db.commit()
    return _get_call(call_id).to_dict()
